import openpyxl
from openpyxl.styles import Font
import os

def registrar_datos(presupuesto, cliente, lugar, fecha, version, tc):
    # Guardar en el mismo directorio del script
    archivo = os.path.join(os.path.dirname(__file__), "Metrados.xlsx")

    if os.path.exists(archivo):
        wb = openpyxl.load_workbook(archivo)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active

        # Crear encabezados de metrado en fila 7
        campos = ["Item", "Descripción", "Und", "Cant", "Largo", "Ancho", "Alto", "Area", "Volumen", "Parcial", "Total"]
        for i, texto in enumerate(campos, start=1):
            celda = ws.cell(row=7, column=i, value=texto)
            celda.font = Font(bold=True)

    # Guardar datos generales
    ws["A1"] = "Presupuesto:"
    ws["B1"] = presupuesto
    ws["A2"] = "Cliente:"
    ws["B2"] = cliente
    ws["A3"] = "Lugar:"
    ws["B3"] = lugar
    ws["A4"] = "Fecha:"
    ws["B4"] = fecha
    ws["A5"] = "Versión:"
    ws["B5"] = version
    ws["A6"] = "TC:"
    ws["B6"] = tc

    wb.save(archivo)
